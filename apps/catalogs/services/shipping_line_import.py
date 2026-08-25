"""Import shipping lines + vessels from Excel (update-or-create by id)."""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from typing import Any

from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from openpyxl import load_workbook

from apps.audit.services.record import record_shipping_line_audit
from apps.catalogs.models import ShippingLine, ShippingLineGroup, Vessel
from apps.catalogs.services.shipping_line_audit import (
    diff_shipping_line_snapshots,
    snapshot_shipping_line,
)
from apps.catalogs.services.shipping_line_export import LINE_SHEET, VESSEL_SHEET


class ShippingLineImportError(Exception):
    """Fatal file-level import error (missing sheets / headers)."""


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_map(headers: list[str]) -> dict[str, int]:
    return {h.strip().lower(): i for i, h in enumerate(headers) if h and str(h).strip()}


def _get(row: list[Any], hmap: dict[str, int], *keys: str) -> str:
    for key in keys:
        idx = hmap.get(key.lower())
        if idx is not None and idx < len(row):
            return _cell_str(row[idx])
    return ""


def _parse_optional_int(raw: str) -> int | None:
    if not raw:
        return None
    try:
        return int(float(raw))
    except (TypeError, ValueError):
        return None


def _parse_bool(raw: str, default: bool = True) -> bool | None:
    if not raw:
        return default
    value = raw.strip().lower()
    if value in ("1", "true", "yes", "y", "si", "sí", "activo", "active"):
        return True
    if value in ("0", "false", "no", "n", "inactivo", "inactive"):
        return False
    return None


def _parse_decimal(raw: str) -> Decimal | None:
    if not raw:
        return None
    try:
        return Decimal(raw.replace(",", "."))
    except (InvalidOperation, ValueError):
        raise ValueError(f"Número inválido: {raw}") from None


def _parse_optional_positive_int(raw: str) -> int | None:
    if not raw:
        return None
    try:
        value = int(float(raw))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Entero inválido: {raw}") from exc
    if value < 0:
        raise ValueError(f"Entero negativo no permitido: {raw}")
    return value


def _load_named_sheet(
    workbook,
    preferred_names: tuple[str, ...],
    fallback_index: int | None,
) -> tuple[dict[str, int], list[tuple[int, list[Any]]]]:
    sheet = None
    lower_names = {n.lower() for n in preferred_names}
    for name in workbook.sheetnames:
        if name.strip().lower() in lower_names:
            sheet = workbook[name]
            break
    if sheet is None and fallback_index is not None and fallback_index < len(
        workbook.sheetnames
    ):
        sheet = workbook[workbook.sheetnames[fallback_index]]
    if sheet is None:
        raise ShippingLineImportError(
            f"No se encontró la hoja «{preferred_names[0]}»."
        )

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ShippingLineImportError(
            f"La hoja «{sheet.title}» está vacía."
        ) from None

    headers = [_cell_str(c) for c in header_row]
    hmap = _header_map(headers)
    if "id" not in hmap:
        raise ShippingLineImportError(
            f"La hoja «{sheet.title}» debe incluir la columna «id» "
            "(vacía = registro nuevo)."
        )

    body: list[tuple[int, list[Any]]] = []
    for excel_row, values in enumerate(rows_iter, start=2):
        body.append((excel_row, list(values) if values is not None else []))
    return hmap, body


def _empty_result_bucket() -> dict[str, Any]:
    return {
        "updated_count": 0,
        "created_count": 0,
        "invalid_count": 0,
        "created": [],
        "invalid": [],
        "updated": [],
    }


def _validation_messages(exc: ValidationError) -> list[str]:
    if hasattr(exc, "message_dict"):
        messages: list[str] = []
        for field, errs in exc.message_dict.items():
            for err in errs:
                messages.append(f"{field}: {err}" if field != "__all__" else str(err))
        return messages or [str(exc)]
    if hasattr(exc, "messages"):
        return [str(m) for m in exc.messages]
    return [str(exc)]


def _resolve_group(
    *,
    group_id_raw: str,
    group_code: str,
) -> tuple[ShippingLineGroup | None, str | None]:
    group_id = _parse_optional_int(group_id_raw)
    if group_id is not None:
        group = ShippingLineGroup.objects.filter(pk=group_id).first()
        if group is None:
            return None, f"group_id {group_id} no existe."
        return group, None
    if group_code:
        group = ShippingLineGroup.objects.filter(code__iexact=group_code).first()
        if group is None:
            return None, f"group_code «{group_code}» no existe."
        return group, None
    return None, "Indica group_id o group_code."


def _upsert_shipping_lines(
    hmap: dict[str, int],
    body: list[tuple[int, list[Any]]],
    *,
    actor,
    request,
) -> tuple[dict[str, Any], dict[str, int]]:
    bucket = _empty_result_bucket()
    created_by_code: dict[str, int] = {}

    for excel_row, cells in body:
        pk_raw = _get(cells, hmap, "id")
        code = _get(cells, hmap, "code", "código", "codigo")
        name = _get(cells, hmap, "name", "nombre")
        if not pk_raw and not code and not name:
            continue

        label = code or name or f"fila {excel_row}"
        errors: list[str] = []

        pk = _parse_optional_int(pk_raw) if pk_raw else None
        if pk_raw and pk is None:
            errors.append(f"id inválido: {pk_raw}")

        group, group_err = _resolve_group(
            group_id_raw=_get(cells, hmap, "group_id"),
            group_code=_get(cells, hmap, "group_code"),
        )
        if group_err:
            errors.append(group_err)

        if not code:
            errors.append("code es obligatorio.")
        if not name:
            errors.append("name es obligatorio.")

        active_raw = _get(cells, hmap, "is_active", "activa", "activo")
        is_active = _parse_bool(active_raw, default=True)
        if is_active is None:
            errors.append(f"is_active inválido: {active_raw}")

        if errors:
            bucket["invalid_count"] += 1
            bucket["invalid"].append(
                {
                    "kind": "shipping_line",
                    "row": excel_row,
                    "id": pk,
                    "code": code,
                    "name": name,
                    "label": label,
                    "errors": errors,
                }
            )
            continue

        try:
            with transaction.atomic():
                if pk is None:
                    if ShippingLine.objects.filter(code__iexact=code).exists():
                        raise ValueError(f"Ya existe una naviera con code «{code}».")
                    line = ShippingLine(
                        code=code,
                        name=name,
                        group=group,
                        is_active=bool(is_active),
                    )
                    line.full_clean()
                    line.save()
                    created_by_code[code.lower()] = line.pk
                    snap = snapshot_shipping_line(
                        ShippingLine.objects.select_related("group").get(pk=line.pk)
                    )
                    record_shipping_line_audit(
                        action="created",
                        summary=f"Creó la naviera {snap['code']} (importación)",
                        shipping_line=line,
                        changes={"created": snap, "source": "excel_import"},
                        actor=actor,
                        request=request,
                        entity=snap,
                    )
                    bucket["created_count"] += 1
                    bucket["created"].append(
                        {
                            "kind": "shipping_line",
                            "row": excel_row,
                            "id": line.pk,
                            "code": line.code,
                            "name": line.name,
                            "label": f"{line.code} — {line.name}",
                        }
                    )
                else:
                    line = (
                        ShippingLine.objects.select_related("group")
                        .filter(pk=pk)
                        .first()
                    )
                    if line is None:
                        raise ValueError(f"Naviera id={pk} no existe.")
                    before = snapshot_shipping_line(line)
                    conflict = (
                        ShippingLine.objects.filter(code__iexact=code)
                        .exclude(pk=pk)
                        .exists()
                    )
                    if conflict:
                        raise ValueError(f"Ya existe otra naviera con code «{code}».")
                    line.code = code
                    line.name = name
                    line.group = group
                    line.is_active = bool(is_active)
                    line.full_clean()
                    line.save()
                    after = snapshot_shipping_line(
                        ShippingLine.objects.select_related("group").get(pk=line.pk)
                    )
                    changes = diff_shipping_line_snapshots(before, after)
                    if changes:
                        changes["source"] = "excel_import"
                        record_shipping_line_audit(
                            action="updated",
                            summary=f"Modificó la naviera {after['code']} (importación)",
                            shipping_line=line,
                            changes=changes,
                            actor=actor,
                            request=request,
                            entity=after,
                        )
                    bucket["updated_count"] += 1
                    bucket["updated"].append(
                        {
                            "kind": "shipping_line",
                            "row": excel_row,
                            "id": line.pk,
                            "code": line.code,
                            "name": line.name,
                            "label": f"{line.code} — {line.name}",
                        }
                    )
        except (ValueError, IntegrityError, ValidationError) as exc:
            msgs = (
                _validation_messages(exc)
                if isinstance(exc, ValidationError)
                else [str(exc)]
            )
            bucket["invalid_count"] += 1
            bucket["invalid"].append(
                {
                    "kind": "shipping_line",
                    "row": excel_row,
                    "id": pk,
                    "code": code,
                    "name": name,
                    "label": label,
                    "errors": msgs,
                }
            )

    return bucket, created_by_code


def _resolve_shipping_line_id(
    *,
    shipping_line_id_raw: str,
    shipping_line_code: str,
    created_by_code: dict[str, int],
) -> tuple[int | None, str | None]:
    line_id = _parse_optional_int(shipping_line_id_raw)
    if line_id is not None:
        if not ShippingLine.objects.filter(pk=line_id).exists():
            # May have been created earlier in this import under a different path.
            return None, f"shipping_line_id {line_id} no existe."
        return line_id, None
    if shipping_line_code:
        mapped = created_by_code.get(shipping_line_code.lower())
        if mapped is not None:
            return mapped, None
        line = ShippingLine.objects.filter(code__iexact=shipping_line_code).first()
        if line is None:
            return None, f"shipping_line_code «{shipping_line_code}» no existe."
        return line.pk, None
    return None, "Indica shipping_line_id o shipping_line_code."


def _apply_vessel_fields(vessel: Vessel, cells: list[Any], hmap: dict[str, int]) -> None:
    vessel.name = _get(cells, hmap, "name", "nombre")
    vessel.ship_code = _get(cells, hmap, "ship_code")
    vessel.vessel_class = _get(cells, hmap, "vessel_class")
    vessel.gross_tonnage = _parse_decimal(_get(cells, hmap, "gross_tonnage"))
    vessel.pax_capacity = _parse_optional_positive_int(_get(cells, hmap, "pax_capacity"))
    vessel.crew_capacity = _parse_optional_positive_int(
        _get(cells, hmap, "crew_capacity")
    )
    vessel.loa_m = _parse_decimal(_get(cells, hmap, "loa_m"))
    vessel.beam_m = _parse_decimal(_get(cells, hmap, "beam_m"))
    vessel.draft_m = _parse_decimal(_get(cells, hmap, "draft_m"))
    vessel.flag = _get(cells, hmap, "flag")
    vessel.year_built = _parse_optional_positive_int(_get(cells, hmap, "year_built"))
    vessel.segment = _get(cells, hmap, "segment")
    vessel.size_category = _get(cells, hmap, "size_category")
    vessel.mooring_line_count = _parse_optional_positive_int(
        _get(cells, hmap, "mooring_line_count")
    )
    vessel.bollard_count = _parse_optional_positive_int(
        _get(cells, hmap, "bollard_count")
    )
    vessel.bollard_swl_t = _parse_decimal(_get(cells, hmap, "bollard_swl_t"))
    active_raw = _get(cells, hmap, "is_active", "activa", "activo")
    is_active = _parse_bool(active_raw, default=True)
    if is_active is None:
        raise ValueError(f"is_active inválido: {active_raw}")
    vessel.is_active = is_active


def _upsert_vessels(
    hmap: dict[str, int],
    body: list[tuple[int, list[Any]]],
    *,
    created_by_code: dict[str, int],
) -> dict[str, Any]:
    bucket = _empty_result_bucket()

    for excel_row, cells in body:
        pk_raw = _get(cells, hmap, "id")
        name = _get(cells, hmap, "name", "nombre")
        if not pk_raw and not name:
            continue

        label = name or f"fila {excel_row}"
        errors: list[str] = []
        pk = _parse_optional_int(pk_raw) if pk_raw else None
        if pk_raw and pk is None:
            errors.append(f"id inválido: {pk_raw}")

        line_id, line_err = _resolve_shipping_line_id(
            shipping_line_id_raw=_get(cells, hmap, "shipping_line_id"),
            shipping_line_code=_get(
                cells, hmap, "shipping_line_code", "naviera_code", "line_code"
            ),
            created_by_code=created_by_code,
        )
        if line_err:
            errors.append(line_err)
        if not name:
            errors.append("name es obligatorio.")

        if errors:
            bucket["invalid_count"] += 1
            bucket["invalid"].append(
                {
                    "kind": "vessel",
                    "row": excel_row,
                    "id": pk,
                    "code": "",
                    "name": name,
                    "label": label,
                    "errors": errors,
                }
            )
            continue

        try:
            with transaction.atomic():
                if pk is None:
                    if Vessel.objects.filter(
                        shipping_line_id=line_id, name__iexact=name
                    ).exists():
                        raise ValueError(
                            f"Ya existe un barco «{name}» en esa naviera."
                        )
                    vessel = Vessel(shipping_line_id=line_id)
                    _apply_vessel_fields(vessel, cells, hmap)
                    vessel.full_clean()
                    vessel.save()
                    bucket["created_count"] += 1
                    bucket["created"].append(
                        {
                            "kind": "vessel",
                            "row": excel_row,
                            "id": vessel.pk,
                            "code": vessel.ship_code or "",
                            "name": vessel.name,
                            "label": vessel.name,
                            "shipping_line_id": vessel.shipping_line_id,
                        }
                    )
                else:
                    vessel = Vessel.objects.filter(pk=pk).first()
                    if vessel is None:
                        raise ValueError(f"Barco id={pk} no existe.")
                    conflict = (
                        Vessel.objects.filter(
                            shipping_line_id=line_id, name__iexact=name
                        )
                        .exclude(pk=pk)
                        .exists()
                    )
                    if conflict:
                        raise ValueError(
                            f"Ya existe otro barco «{name}» en esa naviera."
                        )
                    vessel.shipping_line_id = line_id
                    _apply_vessel_fields(vessel, cells, hmap)
                    vessel.full_clean()
                    vessel.save()
                    bucket["updated_count"] += 1
                    bucket["updated"].append(
                        {
                            "kind": "vessel",
                            "row": excel_row,
                            "id": vessel.pk,
                            "code": vessel.ship_code or "",
                            "name": vessel.name,
                            "label": vessel.name,
                            "shipping_line_id": vessel.shipping_line_id,
                        }
                    )
        except (ValueError, IntegrityError, InvalidOperation, ValidationError) as exc:
            msgs = (
                _validation_messages(exc)
                if isinstance(exc, ValidationError)
                else [str(exc)]
            )
            bucket["invalid_count"] += 1
            bucket["invalid"].append(
                {
                    "kind": "vessel",
                    "row": excel_row,
                    "id": pk,
                    "code": "",
                    "name": name,
                    "label": label,
                    "errors": msgs,
                }
            )

    return bucket


def import_shipping_lines_workbook(file_obj, *, actor=None, request=None) -> dict[str, Any]:
    """
    Update-or-create from Excel with sheets Navieras + Barcos.

    - Empty ``id`` → create
    - Numeric ``id`` → update that PK
    - ``logo`` / timestamps are ignored on import
    """
    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:
        raise ShippingLineImportError(
            "No se pudo leer el Excel. Usa el archivo exportado (.xlsx)."
        ) from exc

    line_hmap, line_body = _load_named_sheet(
        workbook,
        (LINE_SHEET, "shipping_lines", "shipping lines", "lines"),
        fallback_index=0,
    )
    if "code" not in line_hmap and "código" not in line_hmap and "codigo" not in line_hmap:
        raise ShippingLineImportError(
            f"La hoja de navieras debe incluir la columna «code»."
        )
    if "name" not in line_hmap and "nombre" not in line_hmap:
        raise ShippingLineImportError(
            f"La hoja de navieras debe incluir la columna «name»."
        )

    vessel_hmap, vessel_body = _load_named_sheet(
        workbook,
        (VESSEL_SHEET, "vessels", "ships", "barcos"),
        fallback_index=1 if len(workbook.sheetnames) > 1 else None,
    )
    if "name" not in vessel_hmap and "nombre" not in vessel_hmap:
        raise ShippingLineImportError(
            "La hoja de barcos debe incluir la columna «name»."
        )

    lines_bucket, created_by_code = _upsert_shipping_lines(
        line_hmap, line_body, actor=actor, request=request
    )
    vessels_bucket = _upsert_vessels(
        vessel_hmap, vessel_body, created_by_code=created_by_code
    )

    updated_count = lines_bucket["updated_count"] + vessels_bucket["updated_count"]
    created_count = lines_bucket["created_count"] + vessels_bucket["created_count"]
    invalid_count = lines_bucket["invalid_count"] + vessels_bucket["invalid_count"]

    return {
        "updated_count": updated_count,
        "created_count": created_count,
        "invalid_count": invalid_count,
        "shipping_lines": lines_bucket,
        "vessels": vessels_bucket,
        "created": lines_bucket["created"] + vessels_bucket["created"],
        "invalid": lines_bucket["invalid"] + vessels_bucket["invalid"],
    }
