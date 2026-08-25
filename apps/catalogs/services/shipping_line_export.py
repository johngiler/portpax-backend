"""Export shipping lines + vessels to Excel (.xlsx) or CSV zip."""

from __future__ import annotations

import csv
import zipfile
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font

from apps.catalogs.models import ShippingLine, ShippingLineGroup, Vessel

GROUP_SHEET = "Grupos"
LINE_SHEET = "Navieras"
VESSEL_SHEET = "Barcos"

GROUP_HEADERS = [
    "id",
    "code",
    "name",
    "is_active",
]

LINE_HEADERS = [
    "id",
    "code",
    "name",
    "group_id",
    "group_code",
    "group_name",
    "is_active",
]

VESSEL_HEADERS = [
    "id",
    "shipping_line_id",
    "shipping_line_code",
    "name",
    "ship_code",
    "vessel_class",
    "gross_tonnage",
    "pax_capacity",
    "crew_capacity",
    "loa_m",
    "beam_m",
    "draft_m",
    "flag",
    "year_built",
    "segment",
    "size_category",
    "mooring_line_count",
    "bollard_count",
    "bollard_swl_t",
    "is_active",
]


def _fmt_bool(value: bool) -> str:
    return "true" if value else "false"


def _fmt_decimal(value) -> str:
    if value is None:
        return ""
    return str(value)


def group_export_row(group: ShippingLineGroup) -> list:
    return [
        group.pk,
        group.code or "",
        group.name or "",
        _fmt_bool(bool(group.is_active)),
    ]


def shipping_line_export_row(line: ShippingLine) -> list:
    group = getattr(line, "group", None)
    return [
        line.pk,
        line.code or "",
        line.name or "",
        line.group_id or "",
        getattr(group, "code", "") or "",
        getattr(group, "name", "") or "",
        _fmt_bool(bool(line.is_active)),
    ]


def vessel_export_row(vessel: Vessel) -> list:
    line = getattr(vessel, "shipping_line", None)
    return [
        vessel.pk,
        vessel.shipping_line_id or "",
        getattr(line, "code", "") or "",
        vessel.name or "",
        vessel.ship_code or "",
        vessel.vessel_class or "",
        _fmt_decimal(vessel.gross_tonnage),
        vessel.pax_capacity if vessel.pax_capacity is not None else "",
        vessel.crew_capacity if vessel.crew_capacity is not None else "",
        _fmt_decimal(vessel.loa_m),
        _fmt_decimal(vessel.beam_m),
        _fmt_decimal(vessel.draft_m),
        vessel.flag or "",
        vessel.year_built if vessel.year_built is not None else "",
        vessel.segment or "",
        vessel.size_category or "",
        vessel.mooring_line_count if vessel.mooring_line_count is not None else "",
        vessel.bollard_count if vessel.bollard_count is not None else "",
        _fmt_decimal(vessel.bollard_swl_t),
        _fmt_bool(bool(vessel.is_active)),
    ]


def _autosize(sheet) -> None:
    for column_cells in sheet.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        sheet.column_dimensions[column_letter].width = min(max_len + 2, 40)


def iter_export_lines_and_vessels(lines_qs):
    """Yield (lines list, vessels list) from a filtered ShippingLine queryset."""
    lines = list(
        lines_qs.select_related("group").prefetch_related("vessels").order_by("name")
    )
    vessels: list[Vessel] = []
    for line in lines:
        for vessel in line.vessels.all().order_by("name"):
            vessel.shipping_line = line
            vessels.append(vessel)
    return lines, vessels


def iter_export_groups() -> list[ShippingLineGroup]:
    """Full group catalog (reference sheet; independent of line filters)."""
    return list(ShippingLineGroup.objects.order_by("name"))


def build_shipping_lines_xlsx(lines_qs) -> bytes:
    groups = iter_export_groups()
    lines, vessels = iter_export_lines_and_vessels(lines_qs)
    workbook = Workbook()

    group_sheet = workbook.active
    group_sheet.title = GROUP_SHEET
    group_sheet.append(GROUP_HEADERS)
    for cell in group_sheet[1]:
        cell.font = Font(bold=True)
    for group in groups:
        group_sheet.append(group_export_row(group))
    _autosize(group_sheet)

    line_sheet = workbook.create_sheet(LINE_SHEET)
    line_sheet.append(LINE_HEADERS)
    for cell in line_sheet[1]:
        cell.font = Font(bold=True)
    for line in lines:
        line_sheet.append(shipping_line_export_row(line))
    _autosize(line_sheet)

    vessel_sheet = workbook.create_sheet(VESSEL_SHEET)
    vessel_sheet.append(VESSEL_HEADERS)
    for cell in vessel_sheet[1]:
        cell.font = Font(bold=True)
    for vessel in vessels:
        vessel_sheet.append(vessel_export_row(vessel))
    _autosize(vessel_sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _csv_bytes(headers: list[str], rows: list[list]) -> bytes:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def build_shipping_lines_csv_zip(lines_qs) -> bytes:
    """Three CSVs in a zip — Excel is preferred for re-import."""
    groups = iter_export_groups()
    lines, vessels = iter_export_lines_and_vessels(lines_qs)
    group_rows = [group_export_row(group) for group in groups]
    line_rows = [shipping_line_export_row(line) for line in lines]
    vessel_rows = [vessel_export_row(vessel) for vessel in vessels]

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("grupos.csv", _csv_bytes(GROUP_HEADERS, group_rows))
        zf.writestr("navieras.csv", _csv_bytes(LINE_HEADERS, line_rows))
        zf.writestr("barcos.csv", _csv_bytes(VESSEL_HEADERS, vessel_rows))
    return buffer.getvalue()
