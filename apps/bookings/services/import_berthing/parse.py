"""Parse BERTHING PAPERS workbooks into normalized row dicts."""

from __future__ import annotations

import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from apps.bookings.services.import_berthing.aliases import FILE_SPECS, STATUS_MAP


HEADER_ALIASES = {
    "CALL DATE": "call_date",
    "SHIP": "ship",
    "CORP": "corp",
    "BRAND": "brand",
    "CAP/REAL": "pax",
    "BKNG STATUS": "status_raw",
    "STATUS": "status_raw",
    "BOOKING STATUS": "status_raw",
    "ETA": "eta",
    "ETD": "etd",
    "BERTH ASSIG": "berth_assign",
    "BERTH ASSIGN": "berth_assign",
}


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().upper())


def _as_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _as_time(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.time().strftime("%H:%M:%S")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    text = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time().strftime("%H:%M:%S")
        except ValueError:
            continue
    return None


def _as_pax(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = re.sub(r"[^\d]", "", str(value))
    return int(text) if text else None


def map_status(raw: str | None, *, c_means_confirmed: bool) -> str | None:
    if not raw:
        return None
    key = str(raw).strip().upper()
    if key == "C" and c_means_confirmed:
        return "co"
    return STATUS_MAP.get(key)


def _find_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    for idx, row in enumerate(rows):
        col_map: dict[str, int] = {}
        for col, cell in enumerate(row):
            key = _norm_header(cell)
            field = HEADER_ALIASES.get(key)
            if field and field not in col_map:
                col_map[field] = col
        if "call_date" in col_map and "ship" in col_map:
            return idx, col_map
    return None


def parse_workbook(
    path: Path,
    *,
    sheet_name: str,
    port_key: str,
    c_means_confirmed: bool,
) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet «{sheet_name}» not found in {path.name}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    found = _find_header_row(rows)
    if not found:
        raise ValueError(f"No header row in {path.name} / {sheet_name}")
    header_idx, col_map = found

    out: list[dict[str, Any]] = []
    for row_i, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        def cell(field: str) -> Any:
            col = col_map.get(field)
            if col is None or col >= len(row):
                return None
            return row[col]

        ship = str(cell("ship") or "").strip()
        call_date = _as_date(cell("call_date"))
        status_raw = str(cell("status_raw") or "").strip().upper() or None
        brand = str(cell("brand") or "").strip().upper() or None
        corp = str(cell("corp") or "").strip().upper() or None
        berth = str(cell("berth_assign") or "").strip().upper() or None
        if berth in ("", "NA", "N/A", "-", "TBD"):
            berth = None

        # Skip fully empty trailing rows
        if not ship and not call_date and not status_raw and not brand:
            continue

        out.append(
            {
                "source_file": path.name,
                "source_row": row_i,
                "port_key": port_key,
                "call_date": call_date,
                "ship": ship or None,
                "brand": brand,
                "corp": corp,
                "status_raw": status_raw,
                "status": map_status(status_raw, c_means_confirmed=c_means_confirmed),
                "eta": _as_time(cell("eta")),
                "etd": _as_time(cell("etd")),
                "berth_assign": berth,
                "pax": _as_pax(cell("pax")),
            }
        )
    return out


def parse_berthing_folder(folder: Path) -> list[dict[str, Any]]:
    files = {p.name.upper(): p for p in folder.glob("*.xlsx")}
    all_rows: list[dict[str, Any]] = []
    for name_sub, sheet, port_key, c_confirmed in FILE_SPECS:
        match = next(
            (path for uname, path in files.items() if name_sub.upper() in uname),
            None,
        )
        if match is None:
            raise FileNotFoundError(f"Missing workbook matching «{name_sub}» in {folder}")
        all_rows.extend(
            parse_workbook(
                match,
                sheet_name=sheet,
                port_key=port_key,
                c_means_confirmed=c_confirmed,
            )
        )
    return all_rows
