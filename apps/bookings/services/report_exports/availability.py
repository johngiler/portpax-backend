"""Availability chart JSON plus list export (ship / line / port / date / times / position)."""

from __future__ import annotations

import csv
from datetime import date, timedelta
from io import BytesIO, StringIO
from typing import Any

from django.core.files.storage import default_storage
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from apps.bookings.services.report_exports.common import scheduled_bookings_qs
from apps.catalogs.models import Port, Position, PositionComponent
from apps.catalogs.utils.position_code import position_short_code
from apps.bookings.services.validation.conflict_display import (
    booking_conflict_display,
    cell_matches_conflict_filter,
)

# Occupancy density filter (Barcos por día) — exact distinct ships on that day.
SHIPS_PER_DAY_MIN = 1
SHIPS_PER_DAY_MAX = 4


def _related_column_indexes(
    positions: list[Position],
    position_index: dict[int, int],
) -> dict[int, list[int]]:
    """Map each position to chart columns it occupies (self + combined siblings)."""
    ids = [position.id for position in positions]
    related: dict[int, set[int]] = {pid: {pid} for pid in ids}
    links = PositionComponent.objects.filter(
        Q(combined_position_id__in=ids) | Q(source_position_id__in=ids)
    ).values_list("combined_position_id", "source_position_id")
    for combined_id, source_id in links:
        if combined_id in related and source_id in related:
            related[combined_id].add(source_id)
            related[source_id].add(combined_id)
    return {
        pid: sorted(
            position_index[rid] for rid in siblings if rid in position_index
        )
        for pid, siblings in related.items()
    }


def _paged_days_payload(
    *,
    port: Port,
    date_from: date,
    date_to: date,
    columns: list[dict],
    bookings_by_day: dict[date, list[list[dict]]],
    matching_days: list[date],
    page: int | None,
    page_size: int,
    extra: dict | None = None,
) -> dict:
    matched_total = len(matching_days)
    use_page = page if page is not None else 1
    start = (use_page - 1) * page_size
    end = start + page_size
    page_days = matching_days[start:end]
    cell_count = len(columns)
    rows = [
        {
            "date": day.isoformat(),
            "cells": bookings_by_day.get(day, [[] for _ in range(cell_count)]),
        }
        for day in page_days
    ]
    payload = {
        "port_id": port.id,
        "port_code": port.code,
        "port_name": port.name,
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "columns": columns,
        "rows": rows,
        "matched_days": matched_total,
        "page": use_page,
        "page_size": page_size,
        "has_more": end < matched_total,
    }
    if extra:
        payload.update(extra)
    return payload


EXPORT_HEADERS = [
    "Barco",
    "Naviera",
    "Puerto",
    "Fecha",
    "ETA",
    "ETD",
    "Posición",
]


def _format_export_time(value) -> str:
    if value is None:
        return ""
    return value.strftime("%H:%M")


def _port_display_name(port: Port) -> str:
    commercial = (port.commercial_name or "").strip()
    if commercial:
        return f"{port.name} ({commercial})"
    return port.name


def _position_display(booking) -> str:
    if not booking.position_id or booking.position is None:
        return ""
    return position_short_code(booking.port.code, booking.position.code)


def _availability_export_bookings(
    *,
    port_id: int,
    date_from: date,
    date_to: date,
    allowed_ports: set[int] | None = None,
    shipping_line_id: int | None = None,
    vessel_id: int | None = None,
    position_id: int | None = None,
    status: str | None = None,
    statuses: list[str] | None = None,
):
    if allowed_ports is not None and port_id not in allowed_ports:
        raise ValueError("Puerto no permitido.")

    Port.objects.get(pk=port_id)
    status_filters = list(statuses or [])
    if status and status not in status_filters:
        status_filters.append(status)

    qs = scheduled_bookings_qs(
        date_from=date_from,
        date_to=date_to,
        port_id=port_id,
        allowed_ports=allowed_ports,
        shipping_line_id=shipping_line_id,
        vessel_id=vessel_id,
        position_id=position_id,
        status=None,
    )
    if "c" in status_filters:
        from apps.bookings.models import Booking
        from apps.bookings.utils.status_query import apply_booking_status_filters

        base = Booking.objects.filter(
            call_date__gte=date_from,
            call_date__lte=date_to,
            port_id=port_id,
        ).select_related("port", "shipping_line", "vessel", "position")
        if allowed_ports is not None:
            base = base.filter(port_id__in=allowed_ports)
        if shipping_line_id:
            base = base.filter(shipping_line_id=shipping_line_id)
        if vessel_id:
            base = base.filter(vessel_id=vessel_id)
        if position_id:
            base = base.filter(position_id=position_id)
        qs = apply_booking_status_filters(base, status_filters)
    elif status_filters:
        from apps.bookings.utils.status_query import apply_booking_status_filters

        qs = apply_booking_status_filters(qs, status_filters)
    return qs.order_by("call_date", "position__sort_order", "vessel__name")


def _availability_export_rows(
    *,
    port_id: int,
    date_from: date,
    date_to: date,
    allowed_ports: set[int] | None = None,
    shipping_line_id: int | None = None,
    vessel_id: int | None = None,
    position_id: int | None = None,
    status: str | None = None,
    statuses: list[str] | None = None,
) -> tuple[list[str], list[list[str]]]:
    qs = _availability_export_bookings(
        port_id=port_id,
        date_from=date_from,
        date_to=date_to,
        allowed_ports=allowed_ports,
        shipping_line_id=shipping_line_id,
        vessel_id=vessel_id,
        position_id=position_id,
        status=status,
        statuses=statuses,
    )
    rows: list[list[str]] = []
    for booking in qs.iterator(chunk_size=500):
        rows.append(
            [
                booking.vessel.name if booking.vessel_id else "",
                booking.shipping_line.name if booking.shipping_line_id else "",
                _port_display_name(booking.port),
                booking.call_date.strftime("%d/%m/%Y"),
                _format_export_time(booking.eta),
                _format_export_time(booking.etd),
                _position_display(booking),
            ]
        )
    return EXPORT_HEADERS, rows


def _conflict_filter_active(
    *,
    has_conflict: bool | None,
    conflict_severity: str | None,
    conflict_type: str | None,
) -> bool:
    return (
        has_conflict is not None
        or conflict_severity in {"yellow", "red", "green"}
        or bool(conflict_type)
    )


def _soft_focus_active(
    *,
    shipping_line_id: int | None,
    vessel_id: int | None,
    position_id: int | None,
    status_filters: list[str],
    has_conflict: bool | None,
    conflict_severity: str | None,
    conflict_type: str | None,
) -> bool:
    return bool(
        shipping_line_id
        or vessel_id
        or position_id
        or status_filters
        or _conflict_filter_active(
            has_conflict=has_conflict,
            conflict_severity=conflict_severity,
            conflict_type=conflict_type,
        )
    )


def _booking_matches_conflict(
    booking,
    *,
    has_conflict: bool | None,
    conflict_severity: str | None,
    conflict_type: str | None,
) -> bool:
    return cell_matches_conflict_filter(
        {
            "has_conflict": bool(getattr(booking, "has_conflict", False)),
            "conflict_severity": getattr(booking, "conflict_severity", None),
            "conflict_snapshot": getattr(booking, "conflict_snapshot", None),
        },
        has_conflict=has_conflict,
        conflict_severity=conflict_severity,
        conflict_type=conflict_type,
    )


def _focus_match_queryset(
    *,
    port_id: int,
    date_from: date,
    date_to: date,
    allowed_ports: set[int] | None,
    shipping_line_id: int | None,
    vessel_id: int | None,
    position_id: int | None,
    status_filters: list[str],
):
    """Bookings that satisfy soft-focus filters (day discovery only)."""
    from apps.bookings.models import Booking
    from apps.bookings.utils.status_query import apply_booking_status_filters

    if "c" in status_filters:
        qs = Booking.objects.filter(
            call_date__gte=date_from,
            call_date__lte=date_to,
            port_id=port_id,
        ).select_related("port", "shipping_line", "vessel", "position")
        if allowed_ports is not None:
            qs = qs.filter(port_id__in=allowed_ports)
    else:
        qs = scheduled_bookings_qs(
            date_from=date_from,
            date_to=date_to,
            port_id=port_id,
            allowed_ports=allowed_ports,
            status=None,
        )
    if vessel_id:
        qs = qs.filter(vessel_id=vessel_id)
    elif shipping_line_id:
        # Vessel focus supersedes line (same as FE soft-focus match).
        qs = qs.filter(shipping_line_id=shipping_line_id)
    if position_id:
        qs = qs.filter(position_id=position_id)
    if status_filters:
        qs = apply_booking_status_filters(qs, status_filters)
    return qs


def _soft_focus_matching_days(
    *,
    port_id: int,
    date_from: date,
    date_to: date,
    allowed_ports: set[int] | None,
    shipping_line_id: int | None,
    vessel_id: int | None,
    position_id: int | None,
    status_filters: list[str],
    has_conflict: bool | None,
    conflict_severity: str | None,
    conflict_type: str | None,
) -> list[date]:
    qs = _focus_match_queryset(
        port_id=port_id,
        date_from=date_from,
        date_to=date_to,
        allowed_ports=allowed_ports,
        shipping_line_id=shipping_line_id,
        vessel_id=vessel_id,
        position_id=position_id,
        status_filters=status_filters,
    )
    if not _conflict_filter_active(
        has_conflict=has_conflict,
        conflict_severity=conflict_severity,
        conflict_type=conflict_type,
    ):
        return list(
            qs.values_list("call_date", flat=True).distinct().order_by("call_date")
        )

    # Boolean-only conflict filter: stay in SQL (no snapshot walk).
    if (
        has_conflict is not None
        and conflict_severity not in {"yellow", "red", "green"}
        and not conflict_type
    ):
        return list(
            qs.filter(has_conflict=has_conflict)
            .values_list("call_date", flat=True)
            .distinct()
            .order_by("call_date")
        )

    # Severity / type may need snapshot; drop select_related before .only().
    days: set[date] = set()
    conflict_qs = qs.select_related(None).only(
        "call_date",
        "has_conflict",
        "conflict_severity",
        "conflict_snapshot",
    )
    if has_conflict is True:
        conflict_qs = conflict_qs.filter(has_conflict=True)
    elif has_conflict is False:
        conflict_qs = conflict_qs.filter(has_conflict=False)
    for booking in conflict_qs.iterator(chunk_size=500):
        if _booking_matches_conflict(
            booking,
            has_conflict=has_conflict,
            conflict_severity=conflict_severity,
            conflict_type=conflict_type,
        ):
            days.add(booking.call_date)
    return sorted(days)


def _occupied_matching_days(
    *,
    port_id: int,
    date_from: date,
    date_to: date,
    allowed_ports: set[int] | None,
) -> list[date]:
    return list(
        scheduled_bookings_qs(
            date_from=date_from,
            date_to=date_to,
            port_id=port_id,
            allowed_ports=allowed_ports,
            status=None,
        )
        .values_list("call_date", flat=True)
        .distinct()
        .order_by("call_date")
    )


def _density_matching_days(
    *,
    port_id: int,
    date_from: date,
    date_to: date,
    allowed_ports: set[int] | None,
    ships_per_day: int,
) -> list[date]:
    from django.db.models import Count

    rows = (
        scheduled_bookings_qs(
            date_from=date_from,
            date_to=date_to,
            port_id=port_id,
            allowed_ports=allowed_ports,
            status=None,
        )
        .values("call_date")
        .annotate(ship_count=Count("id"))
        .filter(ship_count=ships_per_day)
        .order_by("call_date")
    )
    return [row["call_date"] for row in rows]


def _neighbor_bookings_for_days(
    *,
    port_id: int,
    page_days: list[date],
    allowed_ports: set[int] | None,
    include_cancelled: bool,
):
    """All occupancy (and optional cancelled) calls on the given days — neighbors."""
    if not page_days:
        return []
    day_min = min(page_days)
    day_max = max(page_days)
    bookings = list(
        scheduled_bookings_qs(
            date_from=day_min,
            date_to=day_max,
            port_id=port_id,
            allowed_ports=allowed_ports,
            status=None,
        ).filter(call_date__in=page_days)
    )
    if not include_cancelled:
        return bookings
    seen = {booking.id for booking in bookings}
    for booking in scheduled_bookings_qs(
        date_from=day_min,
        date_to=day_max,
        port_id=port_id,
        allowed_ports=allowed_ports,
        status="c",
    ).filter(call_date__in=page_days):
        if booking.id not in seen:
            bookings.append(booking)
            seen.add(booking.id)
    return bookings


def _availability_columns(port: Port, positions: list[Position]) -> list[dict]:
    return [
        {
            "id": position.id,
            "code": position.code,
            "label": (
                position.code.removeprefix(f"{port.code}-")
                if position.code.startswith(f"{port.code}-")
                else position.code
            ),
            "berth_name": position.berth.name if position.berth_id else "",
            "max_loa_m": (
                str(position.max_loa_m) if position.max_loa_m is not None else None
            ),
        }
        for position in positions
    ]


def _place_bookings_by_day(
    *,
    bookings,
    positions: list[Position],
    columns: list[dict],
    request: Any,
) -> tuple[dict[date, list[list[dict]]], list[dict]]:
    """Fill day → cells grid; may append TBD column when unassigned calls exist."""
    position_index = {position.id: index for index, position in enumerate(positions)}
    has_unassigned = any(
        booking.position_id not in position_index for booking in bookings
    )
    working_columns = list(columns)
    unassigned_index = len(positions) if has_unassigned else None
    if has_unassigned:
        working_columns.append(
            {
                "id": 0,
                "code": "TBD",
                "label": "TBD",
                "berth_name": "Sin asignar",
                "max_loa_m": None,
            }
        )
    cell_count = len(working_columns)
    related_indexes = _related_column_indexes(positions, position_index)
    bookings_by_day: dict[date, list[list[dict]]] = {}
    for booking in bookings:
        if booking.position_id in position_index:
            cell_indexes = related_indexes.get(
                booking.position_id,
                [position_index[booking.position_id]],
            )
        elif unassigned_index is not None:
            cell_indexes = [unassigned_index]
        else:
            continue
        day_cells = bookings_by_day.setdefault(
            booking.call_date,
            [[] for _ in range(cell_count)],
        )
        logo_name = (
            booking.shipping_line.logo.name if booking.shipping_line.logo else None
        )
        logo = default_storage.url(logo_name) if logo_name else None
        if logo and request is not None:
            logo = request.build_absolute_uri(logo)
        vessel_logo_name = booking.vessel.logo.name if booking.vessel.logo else None
        vessel_logo = (
            default_storage.url(vessel_logo_name) if vessel_logo_name else None
        )
        if vessel_logo and request is not None:
            vessel_logo = request.build_absolute_uri(vessel_logo)
        conflict_display = booking_conflict_display(
            has_conflict=bool(getattr(booking, "has_conflict", False)),
            conflict_severity=getattr(booking, "conflict_severity", None),
            snapshot=getattr(booking, "conflict_snapshot", None) or [],
        )
        call = {
            "booking_code": booking.booking_code,
            "status": booking.status,
            "conflict_chips": conflict_display["conflict_chips"],
            "conflict_highlights": conflict_display["conflict_highlights"],
            "position_id": booking.position_id or 0,
            "shipping_line_id": booking.shipping_line_id or 0,
            "shipping_line_name": booking.shipping_line.name,
            "shipping_line_logo": logo,
            "vessel_id": booking.vessel_id or 0,
            "vessel_name": booking.vessel.name,
            "vessel_logo": vessel_logo,
            "loa_m": (
                str(booking.vessel.loa_m)
                if booking.vessel.loa_m is not None
                else None
            ),
            "eta": booking.eta.isoformat() if booking.eta else None,
            "etd": booking.etd.isoformat() if booking.etd else None,
            "actual_pax": booking.actual_pax,
            "planned_pax": booking.planned_pax,
        }
        for cell_index in cell_indexes:
            existing = day_cells[cell_index]
            if any(item.get("booking_code") == call["booking_code"] for item in existing):
                continue
            existing.append(call)
    return bookings_by_day, working_columns


def build_availability_data(
    *,
    port_id: int,
    date_from: date,
    date_to: date,
    allowed_ports: set[int] | None = None,
    shipping_line_id: int | None = None,
    vessel_id: int | None = None,
    position_id: int | None = None,
    status: str | None = None,
    statuses: list[str] | None = None,
    has_conflict: bool | None = None,
    conflict_severity: str | None = None,
    conflict_type: str | None = None,
    ships_per_day: int | None = None,
    occupied_only: bool = False,
    page: int | None = None,
    page_size: int = 30,
    request: Any = None,
) -> dict:
    """JSON payload for the on-screen Availability Chart (day × position).

    Soft-focus filters (vessel/line/position/status/conflict): select matching
    days in the DB, page those days, then load *all* calls on the page days so
    neighbors stay visible. Opacity is a FE concern.
    """
    if allowed_ports is not None and port_id not in allowed_ports:
        raise ValueError("Puerto no permitido.")
    if ships_per_day is not None and (
        ships_per_day < SHIPS_PER_DAY_MIN or ships_per_day > SHIPS_PER_DAY_MAX
    ):
        raise ValueError(
            f"ships_per_day debe estar entre {SHIPS_PER_DAY_MIN} y {SHIPS_PER_DAY_MAX}."
        )
    if page is not None and page < 1:
        raise ValueError("page debe ser >= 1.")
    if page_size < 1 or page_size > 100:
        raise ValueError("page_size debe estar entre 1 y 100.")

    port = Port.objects.get(pk=port_id)
    from apps.catalogs.services.position_combination import exclude_combined_positions

    status_filters = list(statuses or [])
    if status and status not in status_filters:
        status_filters.append(status)
    soft_focus = _soft_focus_active(
        shipping_line_id=shipping_line_id,
        vessel_id=vessel_id,
        position_id=position_id,
        status_filters=status_filters,
        has_conflict=has_conflict,
        conflict_severity=conflict_severity,
        conflict_type=conflict_type,
    )
    paged_mode = soft_focus or occupied_only or ships_per_day is not None

    positions_qs = exclude_combined_positions(
        Position.objects.filter(port_id=port_id, is_active=True)
    )
    # Soft focus keeps all berth columns so same-day neighbors stay visible.
    if position_id and not soft_focus:
        positions_qs = positions_qs.filter(pk=position_id)
    positions = list(
        positions_qs.select_related("berth").order_by("sort_order", "code")
    )
    columns = _availability_columns(port, positions)

    if paged_mode:
        matching_days: list[date]
        if soft_focus:
            matching_days = _soft_focus_matching_days(
                port_id=port_id,
                date_from=date_from,
                date_to=date_to,
                allowed_ports=allowed_ports,
                shipping_line_id=shipping_line_id,
                vessel_id=vessel_id,
                position_id=position_id,
                status_filters=status_filters,
                has_conflict=has_conflict,
                conflict_severity=conflict_severity,
                conflict_type=conflict_type,
            )
            if ships_per_day is not None:
                density_days = set(
                    _density_matching_days(
                        port_id=port_id,
                        date_from=date_from,
                        date_to=date_to,
                        allowed_ports=allowed_ports,
                        ships_per_day=ships_per_day,
                    )
                )
                matching_days = [day for day in matching_days if day in density_days]
        elif ships_per_day is not None:
            matching_days = _density_matching_days(
                port_id=port_id,
                date_from=date_from,
                date_to=date_to,
                allowed_ports=allowed_ports,
                ships_per_day=ships_per_day,
            )
        else:
            matching_days = _occupied_matching_days(
                port_id=port_id,
                date_from=date_from,
                date_to=date_to,
                allowed_ports=allowed_ports,
            )

        use_page = page if page is not None else 1
        start = (use_page - 1) * page_size
        end = start + page_size
        page_days = matching_days[start:end]
        bookings = _neighbor_bookings_for_days(
            port_id=port_id,
            page_days=page_days,
            allowed_ports=allowed_ports,
            include_cancelled="c" in status_filters,
        )
        bookings_by_day, columns = _place_bookings_by_day(
            bookings=bookings,
            positions=positions,
            columns=columns,
            request=request,
        )
        extra = {"ships_per_day": ships_per_day} if ships_per_day is not None else None
        return _paged_days_payload(
            port=port,
            date_from=date_from,
            date_to=date_to,
            columns=columns,
            bookings_by_day=bookings_by_day,
            matching_days=matching_days,
            page=page,
            page_size=page_size,
            extra=extra,
        )

    # Consecutive calendar strip (no sidebar soft-focus / occupancy paging).
    bookings = list(
        scheduled_bookings_qs(
            date_from=date_from,
            date_to=date_to,
            port_id=port_id,
            allowed_ports=allowed_ports,
            status=None,
        )
    )
    bookings_by_day, columns = _place_bookings_by_day(
        bookings=bookings,
        positions=positions,
        columns=columns,
        request=request,
    )
    cell_count = len(columns)
    rows = []
    day = date_from
    while day <= date_to:
        rows.append(
            {
                "date": day.isoformat(),
                "cells": bookings_by_day.get(day, [[] for _ in range(cell_count)]),
            }
        )
        day += timedelta(days=1)

    return {
        "port_id": port.id,
        "port_code": port.code,
        "port_name": port.name,
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "columns": columns,
        "rows": rows,
    }


def build_availability_chart_xlsx(
    *,
    port_id: int,
    date_from: date,
    date_to: date,
    allowed_ports: set[int] | None = None,
    shipping_line_id: int | None = None,
    vessel_id: int | None = None,
    position_id: int | None = None,
    status: str | None = None,
    statuses: list[str] | None = None,
) -> bytes:
    header, rows = _availability_export_rows(
        port_id=port_id,
        date_from=date_from,
        date_to=date_to,
        allowed_ports=allowed_ports,
        shipping_line_id=shipping_line_id,
        vessel_id=vessel_id,
        position_id=position_id,
        status=status,
        statuses=statuses,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Disponibilidad"
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    widths = (28, 28, 24, 12, 10, 10, 14)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_availability_chart_csv(
    *,
    port_id: int,
    date_from: date,
    date_to: date,
    allowed_ports: set[int] | None = None,
    shipping_line_id: int | None = None,
    vessel_id: int | None = None,
    position_id: int | None = None,
    status: str | None = None,
    statuses: list[str] | None = None,
) -> bytes:
    header, rows = _availability_export_rows(
        port_id=port_id,
        date_from=date_from,
        date_to=date_to,
        allowed_ports=allowed_ports,
        shipping_line_id=shipping_line_id,
        vessel_id=vessel_id,
        position_id=position_id,
        status=status,
        statuses=statuses,
    )
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    writer.writerows(rows)
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def availability_filename(
    port_code: str,
    date_from: date,
    date_to: date,
    ext: str = "xlsx",
) -> str:
    return f"availability_{port_code}_{date_from.isoformat()}_{date_to.isoformat()}.{ext}"
