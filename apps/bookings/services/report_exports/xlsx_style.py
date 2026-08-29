"""Shared openpyxl styling for ITM-style operational report exports."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# PortPax / ITM palette (readable in Excel & print).
NAVY = "1B3A5C"
NAVY_MID = "2E5A8A"
SKY = "D6E8F7"
SKY_LIGHT = "EBF4FC"
WHITE = "FFFFFF"
GROWTH_POS = "15803D"
GROWTH_NEG = "DC2626"
MUTED = "64748B"

_THIN = Side(style="thin", color="CBD5E1")
BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

FONT_TITLE = Font(name="Calibri", size=14, bold=True, color=NAVY)
FONT_SECTION = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color=WHITE)
FONT_ROW_LABEL = Font(name="Calibri", size=10, bold=True, color=NAVY)
FONT_DATA = Font(name="Calibri", size=10, color="1E293B")
FONT_TOTAL = Font(name="Calibri", size=10, bold=True, color=NAVY)
FONT_GROWTH_POS = Font(name="Calibri", size=10, bold=True, color=GROWTH_POS)
FONT_GROWTH_NEG = Font(name="Calibri", size=10, bold=True, color=GROWTH_NEG)
FONT_NOTE = Font(name="Calibri", size=9, italic=True, color=MUTED)

FILL_TITLE = PatternFill("solid", fgColor=SKY_LIGHT)
FILL_SECTION = PatternFill("solid", fgColor=NAVY_MID)
FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_ROW_LABEL = PatternFill("solid", fgColor=SKY)
FILL_TOTAL = PatternFill("solid", fgColor=SKY_LIGHT)
FILL_ALT = PatternFill("solid", fgColor=WHITE)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def style_cell(
    cell,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = BORDER_ALL,
    number_format: str | None = None,
) -> None:
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def write_title_row(ws: Worksheet, row: int, title: str, col_span: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    cell = ws.cell(row=row, column=1, value=title)
    style_cell(cell, font=FONT_TITLE, fill=FILL_TITLE, alignment=ALIGN_LEFT, border=None)


def write_section_banner(ws: Worksheet, row: int, label: str, col_span: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    cell = ws.cell(row=row, column=1, value=label)
    style_cell(cell, font=FONT_SECTION, fill=FILL_SECTION, alignment=ALIGN_LEFT)


def write_matrix_header(
    ws: Worksheet,
    row: int,
    *,
    row_label: str,
    month_labels: tuple[str, ...],
    total_label: str = "TOTAL",
) -> None:
    labels = [row_label, *month_labels, total_label]
    for col, text in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        if col == 1:
            style_cell(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_LEFT)
        else:
            style_cell(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)


def write_matrix_row(
    ws: Worksheet,
    row: int,
    *,
    label: str,
    values: list[int | float | None],
    is_total: bool = False,
    alt: bool = False,
) -> None:
    label_font = FONT_TOTAL if is_total else FONT_ROW_LABEL
    label_fill = FILL_TOTAL if is_total else (FILL_ALT if alt else FILL_ROW_LABEL)
    data_font = FONT_TOTAL if is_total else FONT_DATA
    data_fill = FILL_TOTAL if is_total else (FILL_ALT if alt else None)

    label_cell = ws.cell(row=row, column=1, value=label)
    style_cell(label_cell, font=label_font, fill=label_fill, alignment=ALIGN_LEFT)

    for idx, value in enumerate(values, start=2):
        cell = ws.cell(row=row, column=idx, value=value if value else "")
        fmt = "#,##0" if isinstance(value, (int, float)) and value else None
        style_cell(
            cell,
            font=data_font,
            fill=data_fill,
            alignment=ALIGN_RIGHT,
            number_format=fmt,
        )


def write_growth_row(
    ws: Worksheet,
    row: int,
    *,
    label: str,
    values: list[float | None],
) -> None:
    label_cell = ws.cell(row=row, column=1, value=label)
    style_cell(label_cell, font=FONT_ROW_LABEL, fill=FILL_ROW_LABEL, alignment=ALIGN_LEFT)
    for idx, value in enumerate(values, start=2):
        cell = ws.cell(row=row, column=idx)
        if value is None:
            cell.value = ""
            style_cell(cell, font=FONT_DATA, alignment=ALIGN_RIGHT)
            continue
        cell.value = value / 100.0
        font = FONT_GROWTH_POS if value > 0 else FONT_GROWTH_NEG if value < 0 else FONT_DATA
        style_cell(cell, font=font, alignment=ALIGN_RIGHT, number_format="0%")


def autosize_columns(ws: Worksheet, min_width: int = 8, max_width: int = 18) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = max_len
