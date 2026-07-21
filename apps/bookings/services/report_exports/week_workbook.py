"""WEEK-style workbook: movements + monthly ships/pax matrices per port."""

from __future__ import annotations

from collections import defaultdict
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from apps.bookings.services.operational_reports import build_weekly_movements
from apps.bookings.services.report_exports.common import (
    booking_pax,
    scheduled_bookings_qs,
    years_in_range,
)
from apps.catalogs.models import Port


MONTHS = (
    "ENE",
    "FEB",
    "MAR",
    "ABR",
    "MAY",
    "JUN",
    "JUL",
    "AGO",
    "SEP",
    "OCT",
    "NOV",
    "DIC",
)


def build_week_workbook_xlsx(
    *,
    date_from: date,
    date_to: date,
    port_id: int | None = None,
    shipping_line_id: int | None = None,
    allowed_ports: set[int] | None = None,
) -> bytes:
    wb = Workbook()

    # --- Movements (confirmations / cancellations in the window) ---
    movements = build_weekly_movements(
        date_from=date_from,
        date_to=date_to,
        port_id=port_id,
        allowed_ports=allowed_ports,
    )
    ws_m = wb.active
    ws_m.title = "Movimientos"
    ws_m.append(
        [
            "Tipo",
            "Fecha cambio",
            "Código",
            "Call date",
            "Puerto",
            "Naviera",
            "Barco",
            "Estado anterior",
            "Estado nuevo",
        ]
    )
    for cell in ws_m[1]:
        cell.font = Font(bold=True)

    for kind, rows in (
        ("Confirmación", movements.get("confirmations", [])),
        ("Cancelación", movements.get("cancellations", [])),
    ):
        for r in rows:
            ws_m.append(
                [
                    kind,
                    r.get("at", "")[:19].replace("T", " "),
                    r.get("booking_code", ""),
                    r.get("call_date", ""),
                    r.get("port_code", ""),
                    r.get("shipping_line_name", ""),
                    r.get("vessel_name", ""),
                    r.get("from_status", ""),
                    r.get("to_status", ""),
                ]
            )

    # --- Per-port year matrices (ships + pax by month × carrier) ---
    qs = scheduled_bookings_qs(
        date_from=date_from,
        date_to=date_to,
        port_id=port_id,
        shipping_line_id=shipping_line_id,
        allowed_ports=allowed_ports,
    )

    # (port_id, year, line_name, month) -> {ships, pax}
    agg: dict[tuple[int, int, str, int], dict[str, int]] = defaultdict(
        lambda: {"ships": 0, "pax": 0}
    )
    port_meta: dict[int, tuple[str, str]] = {}
    for b in qs.iterator(chunk_size=500):
        port_meta[b.port_id] = (b.port.code, b.port.name)
        line = b.shipping_line.name or b.shipping_line.code or "?"
        key = (b.port_id, b.call_date.year, line, b.call_date.month)
        agg[key]["ships"] += 1
        agg[key]["pax"] += booking_pax(b)

    years = years_in_range(date_from, date_to)
    ports = (
        Port.objects.filter(id__in=port_meta.keys()).order_by("code")
        if port_meta
        else Port.objects.none()
    )

    for port in ports:
        for year in years:
            lines = sorted(
                {
                    line
                    for (pid, y, line, _m) in agg
                    if pid == port.id and y == year
                }
            )
            if not lines:
                continue
            for metric, title_suffix in (("ships", "Ships"), ("pax", "Pax")):
                sheet_name = f"{port.code} {year} {title_suffix}"[:31]
                ws = wb.create_sheet(title=sheet_name)
                ws.append(["Naviera", *MONTHS, "TOTAL"])
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                col_totals = [0] * 12
                for line in lines:
                    row_vals = []
                    row_total = 0
                    for month in range(1, 13):
                        v = agg[(port.id, year, line, month)][metric]
                        row_vals.append(v or "")
                        row_total += v
                        col_totals[month - 1] += v
                    ws.append([line, *row_vals, row_total or ""])
                ws.append(["TOTAL", *[t or "" for t in col_totals], sum(col_totals) or ""])
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def week_workbook_filename(date_from: date, date_to: date) -> str:
    return f"week_report_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
