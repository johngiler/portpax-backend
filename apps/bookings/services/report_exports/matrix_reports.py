"""Year × month matrix reports (port totals, per-port carriers, trends + growth)."""

from __future__ import annotations

from collections import defaultdict
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook

from apps.bookings.services.report_exports.common import (
    booking_pax,
    pax_basis_note,
    scheduled_bookings_qs,
    years_in_range,
    PAX_BASIS_PLANNED,
)
from apps.bookings.services.report_exports.xlsx_style import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    FONT_DATA,
    FONT_HEADER,
    FONT_NOTE,
    FONT_ROW_LABEL,
    FILL_ALT,
    FILL_HEADER,
    FILL_ROW_LABEL,
    autosize_columns,
    style_cell,
    write_growth_row,
    write_matrix_header,
    write_matrix_row,
    write_section_banner,
    write_title_row,
)
from apps.catalogs.models import Port, ShippingLine

MONTH_LABELS = (
    "ENE",
    "FEB",
    "MAR",
    "ABR",
    "MAY",
    "JUN",
    "JUL",
    "AGO",
    "SEP",
    "OCT",
    "NOV",
    "DIC",
)

DEFAULT_MATRIX_SECTION_PAGE_SIZE = 2
DEFAULT_TRENDS_LINE_PAGE_SIZE = 10
MAX_REPORT_PAGE_SIZE = 12


def _paginate_items(
    items: list[Any],
    *,
    page: int | None,
    page_size: int | None,
    default_page_size: int,
) -> tuple[list[Any], dict[str, Any]]:
    total_count = len(items)
    if page is None and page_size is None:
        return items, {
            "page": 1,
            "page_size": total_count,
            "total_count": total_count,
            "has_more": False,
        }
    safe_page = max(1, page or 1)
    safe_size = max(1, min(page_size or default_page_size, MAX_REPORT_PAGE_SIZE))
    start = (safe_page - 1) * safe_size
    end = start + safe_size
    return items[start:end], {
        "page": safe_page,
        "page_size": safe_size,
        "total_count": total_count,
        "has_more": end < total_count,
    }


def _media_url(request, field) -> str | None:
    if not field:
        return None
    try:
        url = field.url
    except ValueError:
        return None
    if request is not None:
        return request.build_absolute_uri(url)
    return url


def _port_logo_map(port_ids: list[int], request=None) -> dict[int, str | None]:
    if not port_ids:
        return {}
    return {
        port.id: _media_url(request, port.logo)
        for port in Port.objects.filter(id__in=port_ids).only("id", "logo")
    }


def _line_logo_map(line_ids: list[int], request=None) -> dict[int, str | None]:
    if not line_ids:
        return {}
    return {
        line.id: _media_url(request, line.logo)
        for line in ShippingLine.objects.filter(id__in=line_ids).only("id", "logo")
    }


def _empty_year_months() -> dict[int, dict[int, dict[str, int]]]:
    return defaultdict(lambda: defaultdict(lambda: {"calls": 0, "pax": 0}))


def _aggregate_by_port(
    qs,
    *,
    pax_basis: str = PAX_BASIS_PLANNED,
) -> tuple[dict[int, dict[int, dict[int, dict[str, int]]]], dict[int, tuple[str, str]]]:
    """port_id -> year -> month -> {calls, pax}."""
    data: dict[int, dict[int, dict[int, dict[str, int]]]] = defaultdict(_empty_year_months)
    meta: dict[int, tuple[str, str]] = {}
    for booking in qs.iterator(chunk_size=500):
        meta[booking.port_id] = (booking.port.code, booking.port.name)
        cell = data[booking.port_id][booking.call_date.year][booking.call_date.month]
        cell["calls"] += 1
        cell["pax"] += booking_pax(booking, pax_basis=pax_basis)
    return data, meta


def _aggregate_by_line(
    qs,
    *,
    pax_basis: str = PAX_BASIS_PLANNED,
) -> tuple[dict[int, dict[int, dict[int, dict[str, int]]]], dict[int, tuple[str, str]]]:
    """shipping_line_id -> year -> month -> {calls, pax}."""
    data: dict[int, dict[int, dict[int, dict[str, int]]]] = defaultdict(_empty_year_months)
    meta: dict[int, tuple[str, str]] = {}
    for booking in qs.iterator(chunk_size=500):
        meta[booking.shipping_line_id] = (
            booking.shipping_line.code,
            booking.shipping_line.name,
        )
        cell = data[booking.shipping_line_id][booking.call_date.year][booking.call_date.month]
        cell["calls"] += 1
        cell["pax"] += booking_pax(booking, pax_basis=pax_basis)
    return data, meta


def _aggregate_trends_by_line(
    qs,
    *,
    pax_basis: str = PAX_BASIS_PLANNED,
) -> tuple[dict[int, dict[int, dict[str, int]]], dict[int, tuple[str, str]]]:
    """shipping_line_id -> year -> {calls, pax}."""
    data: dict[int, dict[int, dict[str, int]]] = defaultdict(
        lambda: defaultdict(lambda: {"calls": 0, "pax": 0})
    )
    meta: dict[int, tuple[str, str]] = {}
    for booking in qs.iterator(chunk_size=500):
        meta[booking.shipping_line_id] = (
            booking.shipping_line.code,
            booking.shipping_line.name,
        )
        cell = data[booking.shipping_line_id][booking.call_date.year]
        cell["calls"] += 1
        cell["pax"] += booking_pax(booking, pax_basis=pax_basis)
    return data, meta


def _year_rows_metric(
    agg: dict[int, dict[int, dict[str, int]]],
    years: list[int],
    metric: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    month_totals = [0] * 12
    grand = 0
    for year in years:
        months = []
        year_total = 0
        for month in range(1, 13):
            value = agg.get(year, {}).get(month, {}).get(metric, 0)
            months.append(value)
            month_totals[month - 1] += value
            year_total += value
        rows.append({"year": year, "months": months, "total": year_total})
        grand += year_total
    rows.append({"year": "total", "months": month_totals, "total": grand, "is_total": True})
    return rows


def _combined_line_month_agg(
    line_data: dict[int, dict[int, dict[int, dict[str, int]]]],
    years: list[int],
) -> dict[int, dict[int, dict[str, int]]]:
    combined: dict[int, dict[int, dict[str, int]]] = defaultdict(
        lambda: defaultdict(lambda: {"calls": 0, "pax": 0})
    )
    for line_agg in line_data.values():
        for year in years:
            for month in range(1, 13):
                cell = line_agg.get(year, {}).get(month, {"calls": 0, "pax": 0})
                combined[year][month]["calls"] += cell["calls"]
                combined[year][month]["pax"] += cell["pax"]
    return combined


def _matrix_section(
    label: str,
    agg: dict[int, dict[int, dict[str, int]]],
    years: list[int],
    *,
    is_total: bool = False,
    logo: str | None = None,
    logo_kind: str | None = None,
) -> dict[str, Any]:
    section: dict[str, Any] = {
        "label": label,
        "calls": _year_rows_metric(agg, years, "calls"),
        "pax": _year_rows_metric(agg, years, "pax"),
        "is_total": is_total,
    }
    if logo_kind:
        section["logo_kind"] = logo_kind
    if logo:
        section["logo"] = logo
    return section


def build_ports_totals_matrix(
    *,
    date_from: date,
    date_to: date,
    without_lta: bool = False,
    pax_basis: str = PAX_BASIS_PLANNED,
    allowed_ports: set[int] | None = None,
    request=None,
    page: int | None = None,
    page_size: int | None = None,
) -> dict[str, Any]:
    qs = scheduled_bookings_qs(
        date_from=date_from,
        date_to=date_to,
        allowed_ports=allowed_ports,
        without_lta=without_lta,
    )
    port_data, port_meta = _aggregate_by_port(qs, pax_basis=pax_basis)
    years = years_in_range(date_from, date_to)

    sections: list[dict[str, Any]] = []
    combined = _combined_line_month_agg(port_data, years)
    sections.append(
        _matrix_section("Total Puertos", combined, years, is_total=True),
    )

    port_ids = sorted(
        port_data.keys(),
        key=lambda pid: (port_meta.get(pid, ("", ""))[0] or "").lower(),
    )
    port_logos = _port_logo_map(port_ids, request)
    for port_id in port_ids:
        code, name = port_meta.get(port_id, ("", f"Puerto {port_id}"))
        sections.append(
            _matrix_section(
                name or code,
                port_data[port_id],
                years,
                logo=port_logos.get(port_id),
                logo_kind="port",
            )
        )

    page_sections, pagination = _paginate_items(
        sections,
        page=page,
        page_size=page_size,
        default_page_size=DEFAULT_MATRIX_SECTION_PAGE_SIZE,
    )

    return {
        "kind": "ports_totals",
        "title": "Bookings totals de puertos",
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "without_lta": without_lta,
        "pax_basis": pax_basis,
        "month_labels": list(MONTH_LABELS),
        "years": years,
        "sections": page_sections,
        "note": pax_basis_note(pax_basis),
        **pagination,
    }


def build_port_carrier_matrix(
    *,
    date_from: date,
    date_to: date,
    port_id: int,
    without_lta: bool = False,
    pax_basis: str = PAX_BASIS_PLANNED,
    allowed_ports: set[int] | None = None,
    request=None,
    page: int | None = None,
    page_size: int | None = None,
) -> dict[str, Any]:
    port = Port.objects.get(pk=port_id)
    qs = scheduled_bookings_qs(
        date_from=date_from,
        date_to=date_to,
        port_id=port_id,
        allowed_ports=allowed_ports,
        without_lta=without_lta,
    )
    line_data, line_meta = _aggregate_by_line(qs, pax_basis=pax_basis)
    years = years_in_range(date_from, date_to)

    sections: list[dict[str, Any]] = []
    combined = _combined_line_month_agg(line_data, years) if line_data else {}
    if not combined:
        combined = defaultdict(lambda: defaultdict(lambda: {"calls": 0, "pax": 0}))
    port_logo = _media_url(request, port.logo)
    sections.append(
        _matrix_section(
            f"Total {port.name}",
            combined,
            years,
            is_total=True,
            logo=port_logo,
            logo_kind="port",
        )
    )

    line_ids = sorted(
        line_data.keys(),
        key=lambda lid: (line_meta.get(lid, ("", ""))[1] or "").lower(),
    )
    line_logos = _line_logo_map(line_ids, request)
    for line_id in line_ids:
        code, name = line_meta.get(line_id, ("", f"Línea {line_id}"))
        sections.append(
            _matrix_section(
                name or code,
                line_data[line_id],
                years,
                logo=line_logos.get(line_id),
                logo_kind="shipping_line",
            )
        )

    page_sections, pagination = _paginate_items(
        sections,
        page=page,
        page_size=page_size,
        default_page_size=DEFAULT_MATRIX_SECTION_PAGE_SIZE,
    )

    return {
        "kind": "port_carrier",
        "title": f"Bookings totals por puerto — {port.name}",
        "port": {
            "id": port.id,
            "code": port.code,
            "name": port.name,
            "logo": port_logo,
        },
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "without_lta": without_lta,
        "pax_basis": pax_basis,
        "month_labels": list(MONTH_LABELS),
        "years": years,
        "sections": page_sections,
        "note": pax_basis_note(pax_basis),
        **pagination,
    }


def _growth_pct(current: int, previous: int) -> float | None:
    if previous <= 0:
        return None if current <= 0 else 100.0
    return round(((current - previous) / previous) * 100)


def build_port_trends(
    *,
    date_from: date,
    date_to: date,
    port_id: int,
    without_lta: bool = False,
    pax_basis: str = PAX_BASIS_PLANNED,
    allowed_ports: set[int] | None = None,
    request=None,
    page: int | None = None,
    page_size: int | None = None,
) -> dict[str, Any]:
    port = Port.objects.get(pk=port_id)
    qs = scheduled_bookings_qs(
        date_from=date_from,
        date_to=date_to,
        port_id=port_id,
        allowed_ports=allowed_ports,
        without_lta=without_lta,
    )
    line_data, line_meta = _aggregate_trends_by_line(qs, pax_basis=pax_basis)
    years = years_in_range(date_from, date_to)

    lines: list[dict[str, Any]] = []
    line_ids = sorted(
        line_data.keys(),
        key=lambda lid: (line_meta.get(lid, ("", ""))[1] or "").lower(),
    )
    line_logos = _line_logo_map(line_ids, request)
    for line_id in line_ids:
        code, name = line_meta.get(line_id, ("", f"Línea {line_id}"))
        by_year = []
        total_ships = 0
        total_pax = 0
        for year in years:
            cell = line_data[line_id].get(year, {"calls": 0, "pax": 0})
            by_year.append(
                {
                    "year": year,
                    "ships": cell["calls"],
                    "pax": cell["pax"],
                }
            )
            total_ships += cell["calls"]
            total_pax += cell["pax"]

        growth: list[dict[str, Any]] = []
        for idx, year in enumerate(years):
            pax = by_year[idx]["pax"]
            prev = by_year[idx - 1]["pax"] if idx > 0 else 0
            growth.append(
                {
                    "year": year,
                    "pct": _growth_pct(pax, prev) if idx > 0 else None,
                }
            )

        lines.append(
            {
                "shipping_line_id": line_id,
                "code": code,
                "name": name,
                "logo": line_logos.get(line_id),
                "by_year": by_year,
                "growth": growth,
                "total_ships": total_ships,
                "total_pax": total_pax,
            }
        )

    port_logo = _media_url(request, port.logo)
    page_lines, pagination = _paginate_items(
        lines,
        page=page,
        page_size=page_size,
        default_page_size=DEFAULT_TRENDS_LINE_PAGE_SIZE,
    )
    return {
        "kind": "port_trends",
        "title": f"Trends por puerto — {port.name}",
        "port": {
            "id": port.id,
            "code": port.code,
            "name": port.name,
            "logo": port_logo,
        },
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "without_lta": without_lta,
        "pax_basis": pax_basis,
        "years": years,
        "lines": page_lines,
        "note": (
            f"{pax_basis_note(pax_basis)} Growth % = variación YoY de PAX."
        ),
        **pagination,
    }


def _write_matrix_block(
    ws,
    start_row: int,
    *,
    title: str,
    sections: list[dict[str, Any]],
    metric_key: str,
    row_label_header: str,
) -> int:
    col_span = 14
    write_title_row(ws, start_row, title, col_span)
    row = start_row + 2
    for section in sections:
        write_section_banner(ws, row, section["label"], col_span)
        row += 1
        write_matrix_header(ws, row, row_label=row_label_header, month_labels=MONTH_LABELS)
        row += 1
        rows = section[metric_key]
        for idx, data_row in enumerate(rows):
            label = str(data_row["year"]) if data_row["year"] != "total" else "TOTAL"
            write_matrix_row(
                ws,
                row,
                label=label,
                values=data_row["months"] + [data_row["total"]],
                is_total=data_row.get("is_total", False),
                alt=idx % 2 == 1 and not data_row.get("is_total"),
            )
            row += 1
        row += 1
    return row


def _write_dual_matrix_sheet(
    wb: Workbook,
    *,
    sheet_title: str,
    report: dict[str, Any],
    calls_title: str,
    pax_title: str,
) -> None:
    ws = wb.active if wb.sheetnames == ["Sheet"] else wb.create_sheet(sheet_title)
    ws.title = sheet_title[:31]
    if wb.sheetnames[0] == "Sheet" and ws.title != "Sheet":
        wb.remove(wb["Sheet"])

    note_row = 1
    note = report.get("note", "")
    if report.get("without_lta"):
        note = f"{note} Sin LTA / CL / LTD." if note else "Sin LTA / CL / LTD."
    if note:
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=14)
        cell = ws.cell(row=note_row, column=1, value=note)
        style_cell(cell, font=FONT_NOTE, border=None)

    start = note_row + 1
    end_calls = _write_matrix_block(
        ws,
        start,
        title=calls_title,
        sections=report["sections"],
        metric_key="calls",
        row_label_header="AÑO",
    )
    _write_matrix_block(
        ws,
        end_calls + 1,
        title=pax_title,
        sections=report["sections"],
        metric_key="pax",
        row_label_header="AÑO",
    )
    autosize_columns(ws)


def build_ports_totals_matrix_xlsx(**kwargs) -> bytes:
    report = build_ports_totals_matrix(**kwargs)
    wb = Workbook()
    _write_dual_matrix_sheet(
        wb,
        sheet_title="Totals Puertos",
        report=report,
        calls_title="CALL SUMMARY ITM PORTS",
        pax_title="PASSENGER SUMMARY ITM PORTS",
    )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_port_carrier_matrix_xlsx(**kwargs) -> bytes:
    report = build_port_carrier_matrix(**kwargs)
    port_name = report["port"]["name"]
    wb = Workbook()
    _write_dual_matrix_sheet(
        wb,
        sheet_title=report["port"]["code"][:20],
        report=report,
        calls_title=f"CALL SUMMARY {port_name.upper()}",
        pax_title=f"PASSENGER SUMMARY {port_name.upper()}",
    )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_port_trends_xlsx(**kwargs) -> bytes:
    report = build_port_trends(**kwargs)
    years = report["years"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Trends"
    port_name = report["port"]["name"]

    write_title_row(ws, 1, f"TRENDS — {port_name.upper()}", 2 + len(years) * 2)
    row = 3
    header = ["Naviera"]
    for year in years:
        header.extend([f"{year} SHIPS", f"{year} PAX"])
    header.extend(["Total SHIPS", "Total PAX"])
    for col, text in enumerate(header, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        style_cell(
            cell,
            font=FONT_HEADER,
            fill=FILL_HEADER,
            alignment=ALIGN_CENTER if col > 1 else ALIGN_LEFT,
        )
    row += 1

    for idx, line in enumerate(report["lines"]):
        values: list[Any] = [line["name"]]
        for cell in line["by_year"]:
            values.extend([cell["ships"] or "", cell["pax"] or ""])
        values.extend([line["total_ships"] or "", line["total_pax"] or ""])
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value if value != 0 else "")
            font = FONT_ROW_LABEL if col == 1 else FONT_DATA
            fill = FILL_ROW_LABEL if col == 1 else (FILL_ALT if idx % 2 else None)
            style_cell(
                cell,
                font=font,
                fill=fill,
                alignment=ALIGN_LEFT if col == 1 else ALIGN_RIGHT,
                number_format="#,##0" if isinstance(value, int) and col > 1 else None,
            )
        row += 1

    row += 1
    write_title_row(ws, row, "GROWTH PERCENTAGE (PAX YoY)", 1 + len(years))
    row += 2
    growth_header = ["Naviera", *[str(y) for y in years]]
    for col, text in enumerate(growth_header, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        style_cell(
            cell,
            font=FONT_HEADER,
            fill=FILL_HEADER,
            alignment=ALIGN_CENTER if col > 1 else ALIGN_LEFT,
        )
    row += 1

    for line in report["lines"]:
        pct_values = [g["pct"] for g in line["growth"]]
        write_growth_row(ws, row, label=line["name"], values=pct_values)
        row += 1

    autosize_columns(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def ports_totals_matrix_filename(date_from: date, date_to: date) -> str:
    return f"totals_puertos_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"


def port_carrier_matrix_filename(
    port_code: str,
    date_from: date,
    date_to: date,
) -> str:
    safe = (port_code or "port").replace(" ", "_")
    return f"totals_{safe}_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"


def port_trends_filename(port_code: str, date_from: date, date_to: date) -> str:
    safe = (port_code or "port").replace(" ", "_")
    return f"trends_{safe}_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
