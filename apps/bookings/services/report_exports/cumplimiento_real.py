"""Cumplimiento REAL-only — PAX by shipping line with share % (no projection)."""

from __future__ import annotations

import csv
from collections import defaultdict
from datetime import date
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font

from apps.bookings.services.report_exports.common import (
    booking_pax,
    scheduled_bookings_qs,
    years_in_range,
)


def build_cumplimiento_real(
    *,
    date_from: date,
    date_to: date,
    port_id: int | None = None,
    allowed_ports: set[int] | None = None,
) -> dict:
    qs = scheduled_bookings_qs(
        date_from=date_from,
        date_to=date_to,
        port_id=port_id,
        allowed_ports=allowed_ports,
    )

    # line_id -> year -> pax
    by_line_year: dict[int, dict[int, int]] = defaultdict(lambda: defaultdict(int))
    line_meta: dict[int, dict] = {}
    year_totals: dict[int, int] = defaultdict(int)

    for b in qs.iterator(chunk_size=500):
        pax = booking_pax(b)
        by_line_year[b.shipping_line_id][b.call_date.year] += pax
        year_totals[b.call_date.year] += pax
        line_meta[b.shipping_line_id] = {
            "shipping_line_id": b.shipping_line_id,
            "code": b.shipping_line.code,
            "name": b.shipping_line.name,
        }

    years = years_in_range(date_from, date_to)
    lines = sorted(line_meta.values(), key=lambda x: x["name"].lower())

    rows = []
    for meta in lines:
        lid = meta["shipping_line_id"]
        by_year = []
        total = 0
        for y in years:
            pax = by_line_year[lid][y]
            yt = year_totals[y]
            pct = round((pax / yt) * 100, 1) if yt else 0.0
            by_year.append({"year": y, "pax": pax, "share_pct": pct})
            total += pax
        rows.append({**meta, "by_year": by_year, "total_pax": total})

    grand = sum(year_totals[y] for y in years)
    return {
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "years": years,
        "year_totals": [{"year": y, "pax": year_totals[y]} for y in years],
        "grand_total_pax": grand,
        "lines": rows,
        "note": (
            "Solo PAX real/planificado de bookings (sin fila de proyección anual). "
            "% = participación sobre el total del año."
        ),
    }


def _cumplimiento_real_table(data: dict) -> tuple[list, list[list]]:
    years = data["years"]
    header = ["Naviera", "Código"]
    for y in years:
        header.extend([f"{y} PAX", f"{y} %"])
    header.append("Total PAX")
    rows: list[list] = []
    for row in data["lines"]:
        out = [row["name"], row["code"]]
        for cell in row["by_year"]:
            out.extend([cell["pax"] or "", cell["share_pct"] if cell["pax"] else ""])
        out.append(row["total_pax"] or "")
        rows.append(out)
    total_row = ["TOTAL", ""]
    for yt in data["year_totals"]:
        total_row.extend([yt["pax"] or "", 100 if yt["pax"] else ""])
    total_row.append(data["grand_total_pax"] or "")
    rows.append(total_row)
    return header, rows


def build_cumplimiento_real_xlsx(**kwargs) -> bytes:
    data = build_cumplimiento_real(**kwargs)
    header, rows = _cumplimiento_real_table(data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cumplimiento REAL"
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_cumplimiento_real_csv(**kwargs) -> bytes:
    data = build_cumplimiento_real(**kwargs)
    header, rows = _cumplimiento_real_table(data)
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    writer.writerows(rows)
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def cumplimiento_real_filename(
    date_from: date,
    date_to: date,
    ext: str = "xlsx",
) -> str:
    return f"cumplimiento_real_{date_from.isoformat()}_{date_to.isoformat()}.{ext}"
