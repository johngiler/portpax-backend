"""Parse ITM mass-booking Excel or pasted TSV (Ship, Port, Arrival, Departure, …)."""

from __future__ import annotations

from datetime import date, datetime, time
from typing import Any, Iterable

from openpyxl import load_workbook

from apps.bookings.services.import_mass.parse_dates import parse_flexible_datetime

REQUIRED_HEADERS = ("Ship", "Port", "Arrival", "Departure")


class ItmParseError(Exception):
    pass


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _as_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date) and not isinstance(value, datetime):
        return datetime.combine(value, time(0, 0))
    return parse_flexible_datetime(value)


def _parse_itm_table(
    headers: list[str],
    body_rows: Iterable[tuple[int, list[Any]]],
) -> list[dict[str, Any]]:
    lower_map = {h.lower(): i for i, h in enumerate(headers) if h}
    for required in REQUIRED_HEADERS:
        if required.lower() not in lower_map:
            raise ItmParseError(
                f"Falta la columna «{required}». "
                "Formato esperado: Ship, Port, Arrival, Departure, …"
            )

    ship_i = lower_map["ship"]
    port_i = lower_map["port"]
    arr_i = lower_map["arrival"]
    dep_i = lower_map["departure"]
    vendor_i = lower_map.get("vendor name")
    call_type_i = lower_map.get("call type")
    position_i = None
    for key in (
        "position",
        "posición",
        "posicion",
        "position code",
        "berth",
        "pos",
    ):
        if key in lower_map:
            position_i = lower_map[key]
            break

    parsed: list[dict[str, Any]] = []
    for excel_row, values in body_rows:
        ship = _cell_str(values[ship_i] if ship_i < len(values) else None)
        port = _cell_str(values[port_i] if port_i < len(values) else None)
        if not ship and not port:
            continue

        arrival = _as_datetime(values[arr_i] if arr_i < len(values) else None)
        departure = _as_datetime(values[dep_i] if dep_i < len(values) else None)
        vendor = (
            _cell_str(values[vendor_i] if vendor_i is not None and vendor_i < len(values) else None)
            if vendor_i is not None
            else ""
        )
        call_type = (
            _cell_str(
                values[call_type_i]
                if call_type_i is not None and call_type_i < len(values)
                else None
            )
            if call_type_i is not None
            else ""
        )
        position_raw = (
            _cell_str(
                values[position_i]
                if position_i is not None and position_i < len(values)
                else None
            )
            if position_i is not None
            else ""
        )

        parsed.append(
            {
                "row_number": excel_row,
                "ship": ship,
                "port_raw": port,
                "arrival": arrival,
                "departure": departure,
                "vendor_name": vendor,
                "call_type": call_type,
                "position_raw": position_raw,
            }
        )

    if not parsed:
        raise ItmParseError("No se encontraron filas de reservas.")
    return parsed


def parse_itm_workbook(file_obj) -> list[dict[str, Any]]:
    """Return raw rows from the first sheet of an ITM-format xlsx."""
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration as exc:
            raise ItmParseError("El archivo está vacío.") from exc

        headers = [_cell_str(h) for h in (header or ())]
        body: list[tuple[int, list[Any]]] = []
        for excel_row, row in enumerate(rows_iter, start=2):
            if row is None:
                continue
            body.append((excel_row, list(row)))
        return _parse_itm_table(headers, body)
    finally:
        wb.close()


def parse_itm_tsv(text: str) -> list[dict[str, Any]]:
    """Parse clipboard paste from Excel/Sheets/email as ITM columns."""
    raw = (text or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not raw:
        raise ItmParseError(
            "Pega al menos una fila con Ship, Port, Arrival y Departure."
        )

    lines = [ln for ln in raw.split("\n") if ln.strip()]
    if not lines:
        raise ItmParseError(
            "Pega al menos una fila con Ship, Port, Arrival y Departure."
        )

    vertical = _reshape_vertical_itm_lines(lines)
    if vertical is not None:
        headers, body_rows = vertical
        body = [(i, row) for i, row in enumerate(body_rows, start=2)]
        return _parse_itm_table(headers, body)

    def split_line(line: str) -> list[str]:
        if "\t" in line:
            return [_cell_str(c) for c in line.split("\t")]
        if ";" in line:
            return [_cell_str(c) for c in line.split(";")]
        return [_cell_str(line)]

    headers = split_line(lines[0])
    body = [(i, split_line(line)) for i, line in enumerate(lines[1:], start=2)]
    if not body:
        raise ItmParseError(
            "Incluye la fila de encabezados (Ship, Port, Arrival, Departure) "
            "y al menos una fila de datos."
        )
    return _parse_itm_table(headers, body)


_VERTICAL_ITM_HEADERS = (
    "Ship",
    "Port",
    "Arrival",
    "Departure",
    "Vendor Name",
    "Call Type",
    "Position",
)
_VERTICAL_ITM_KEYS = {h.lower(): h for h in _VERTICAL_ITM_HEADERS}
# Spanish / aliases for vertical email paste header detection.
_VERTICAL_ITM_KEYS.update(
    {
        "posición": "Position",
        "posicion": "Position",
        "position code": "Position",
        "berth": "Position",
        "pos": "Position",
    }
)


def _reshape_vertical_itm_lines(
    lines: list[str],
) -> tuple[list[str], list[list[str]]] | None:
    """
    Outlook/email often copies ITM tables as one field per line:
    headers, then repeating value blocks of the same width.
    """
    trimmed = [_cell_str(ln) for ln in lines if _cell_str(ln)]
    if len(trimmed) < 8:
        return None

    mostly_single = (
        sum(1 for ln in trimmed if "\t" not in ln and ";" not in ln)
        >= len(trimmed) * 0.85
    )
    if not mostly_single:
        return None

    headers: list[str] = []
    for line in trimmed:
        key = line.lower()
        if key not in _VERTICAL_ITM_KEYS:
            break
        headers.append(_VERTICAL_ITM_KEYS[key])

    if len(headers) < 4:
        return None
    header_keys = {h.lower() for h in headers}
    for required in REQUIRED_HEADERS:
        if required.lower() not in header_keys:
            return None

    width = len(headers)
    data = trimmed[len(headers) :]
    if len(data) < width or len(data) % width != 0:
        return None

    body = [data[i : i + width] for i in range(0, len(data), width)]
    return headers, body

