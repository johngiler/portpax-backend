"""Export mass-import retry / pending rows to Excel (.xlsx)."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HEADERS = [
    "Ship",
    "Port",
    "Arrival",
    "Departure",
    "Vendor Name",
    "Call Type",
    "Issues",
]


def _cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _arrival_departure(row: dict[str, Any], which: str) -> str:
    """Prefer ITM-style datetime; fall back to date + time parts."""
    call_date = _cell(row.get("call_date"))
    time_key = "eta" if which == "arrival" else "etd"
    time_val = _cell(row.get(time_key))
    if not call_date:
        return ""
    if time_val:
        return f"{call_date} {time_val[:5]}"
    return call_date


def build_import_rows_xlsx(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pendientes"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        if not isinstance(row, dict):
            continue
        issues = row.get("issues") or []
        if isinstance(issues, list):
            issues_text = "; ".join(str(i) for i in issues if i)
        else:
            issues_text = str(issues)
        ws.append(
            [
                _cell(row.get("ship") or row.get("vessel_name")),
                _cell(row.get("port_raw") or row.get("port_name") or row.get("port")),
                _arrival_departure(row, "arrival"),
                _arrival_departure(row, "departure"),
                _cell(row.get("vendor_name")),
                _cell(row.get("call_type") or "Standard"),
                issues_text,
            ]
        )

    for idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    ws.column_dimensions["G"].width = 40

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
