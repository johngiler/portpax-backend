"""Parse Excel or pasted dates for availability filter (dates only)."""

from __future__ import annotations

from datetime import date
from typing import Any, Iterable

from openpyxl import load_workbook

from apps.bookings.services.import_mass.parse_dates import parse_flexible_date
from apps.bookings.services.import_mass.parse_itm import ItmParseError

DATE_ALIASES = (
    "fecha",
    "fechas",
    "date",
    "dates",
    "call_date",
    "arrival",
    "arribo",
)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _as_date(value: Any) -> date | None:
    return parse_flexible_date(value)


def _find_date_col(headers: list[str]) -> int | None:
    lower_map = {h.lower(): i for i, h in enumerate(headers) if h}
    for alias in DATE_ALIASES:
        if alias in lower_map:
            return lower_map[alias]
    return None


def _header_looks_like_labels(cells: list[Any]) -> bool:
    return any(_cell_str(c).lower() in DATE_ALIASES for c in cells)


def _build_payload_from_date_values(
    dated_rows: Iterable[tuple[int, date]],
) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    dates: set[str] = set()

    for row_number, call_date in dated_rows:
        iso = call_date.isoformat()
        dates.add(iso)
        rows.append({"row_number": row_number, "call_date": iso})

    if not dates:
        raise ItmParseError(
            "No se encontraron fechas válidas. "
            "Pega o sube solo la columna de fechas (una por fila)."
        )

    sorted_dates = sorted(dates)
    return {
        "rows": rows,
        "dates": sorted_dates,
        "date_from": sorted_dates[0],
        "date_to": sorted_dates[-1],
        "total": len(rows),
    }


def _iter_dates_from_matrix(
    body_rows: Iterable[tuple[int, list[Any]]],
    *,
    date_col: int | None,
) -> list[tuple[int, date]]:
    """Extract dates from a date column, or from any cell when date_col is None."""
    found: list[tuple[int, date]] = []
    for row_number, values in body_rows:
        if date_col is not None:
            call_date = _as_date(values[date_col] if date_col < len(values) else None)
            if call_date is not None:
                found.append((row_number, call_date))
            continue
        for cell in values:
            call_date = _as_date(cell)
            if call_date is not None:
                found.append((row_number, call_date))
                break
    return found


def parse_availability_list_workbook(file_obj) -> dict[str, Any]:
    """Parse a workbook; uses the fechas column when present, otherwise any date cells."""
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            first = next(rows_iter)
        except StopIteration as exc:
            raise ItmParseError("El archivo está vacío.") from exc

        first_cells = list(first or ())
        body: list[tuple[int, list[Any]]] = []

        if _header_looks_like_labels(first_cells):
            headers = [_cell_str(h) for h in first_cells]
            date_col = _find_date_col(headers)
            if date_col is None:
                raise ItmParseError(
                    "Falta la columna «fecha» / «fechas». "
                    "Puedes subir solo esa columna."
                )
            for excel_row, row in enumerate(rows_iter, start=2):
                if row is None:
                    continue
                body.append((excel_row, list(row)))
            dated = _iter_dates_from_matrix(body, date_col=date_col)
        else:
            # First row is data (typically a single column of dates).
            body.append((1, list(first_cells)))
            for excel_row, row in enumerate(rows_iter, start=2):
                if row is None:
                    continue
                body.append((excel_row, list(row)))
            dated = _iter_dates_from_matrix(body, date_col=None)

        return _build_payload_from_date_values(dated)
    finally:
        wb.close()


def parse_availability_list_tsv(text: str) -> dict[str, Any]:
    """
    Parse clipboard paste of dates (one per line).

    Accepts optional header (fecha / fechas), tab/semicolon columns
    (only the date cells are used), or a raw list of dates.
    """
    raw = (text or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not raw:
        raise ItmParseError("Pega al menos una fecha.")

    lines = [ln for ln in raw.split("\n") if ln.strip()]
    if not lines:
        raise ItmParseError("Pega al menos una fecha.")

    def split_line(line: str) -> list[str]:
        if "\t" in line:
            return [_cell_str(c) for c in line.split("\t")]
        if ";" in line:
            return [_cell_str(c) for c in line.split(";")]
        return [_cell_str(line)]

    first_cells = split_line(lines[0])

    if _header_looks_like_labels(first_cells):
        headers = first_cells
        date_col = _find_date_col(headers)
        if date_col is None:
            raise ItmParseError(
                "Falta la columna «fecha» / «fechas» en el encabezado."
            )
        body = [(i, split_line(line)) for i, line in enumerate(lines[1:], start=2)]
        dated = _iter_dates_from_matrix(body, date_col=date_col)
        return _build_payload_from_date_values(dated)

    body = [(i, split_line(line)) for i, line in enumerate(lines, start=1)]
    dated = _iter_dates_from_matrix(body, date_col=None)
    return _build_payload_from_date_values(dated)
